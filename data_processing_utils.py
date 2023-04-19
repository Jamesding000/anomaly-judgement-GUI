import pandas as pd
import numpy as np
import openpyxl
import csv
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

CURRENT_LINE_NO_FILE_NAME = 'data/currentLineNo.csv'
CHECKED_MC_FILE_NAME = 'data/毛刺判断(GUI自动填写).xlsx'
SHEET_NAME = 'Sheet1'

kp_df = pd.read_excel('data/平台画像数据库.xlsx',sheet_name='知识词典').dropna(subset=['相关平台'])

cleaned_basic_info = pd.read_csv("data/平台基本信息(452个+已清洗).csv")
cleaned_basic_info = cleaned_basic_info.set_index('FullName')

mc_df = pd.read_excel("data/毛刺机器校对(全部9162).xlsx")
cleaned_mc_info = mc_df.drop(columns=['是否确实为毛刺？','平台编号','平台','如何修改','证据/依据','备注'])

checked_mc_df = pd.read_excel('data/毛刺判断(GUI自动填写).xlsx').dropna(how='all')

def load_dataset():
    # 导入需要的数据
    t1 = np.load(r'data/data_3darray.npy')
    calendar = pd.read_csv(r'data/standard_calendar.csv')
    platforms = pd.read_csv(r'data/standard_platform.csv')
    features = pd.read_csv(r'data/feature.csv')
    static_data = pd.read_csv('data/1031个平台基本信息.csv')
    mc = pd.read_excel(r'data/疑似毛刺数据列表.xlsx')
    mc = mc.set_index("时间序列名称")
    t1 = t1.astype(np.float32)

    # 分别设置1、2、3维的标签
    index1 = pd.Series(["TradingVolume","AveReturn","InvestorNum","AveLimTime","LoanNum","CumulateRepay","F30Repay","F60Repay","RegistedCapital","PlatformBackground","Unknown1","Unknown2"])
    index2 = pd.Series(platforms.values.flatten())
    index3 = pd.Series(calendar.values.flatten())

    # 展开三维数组
    value = t1.flatten()
    value = pd.Series(value)
    value.name = 'value'

    # 展开标签
    import itertools

    # index的笛卡尔乘积。注意高维在前，低维在后。
    prod = itertools.product(index3,index2,index1)

    # 转换成DataFrame
    prod = pd.DataFrame([x for x in prod])
    prod.columns = ["日期","平台","特征"]

    # 合并成一个DataFrame
    df = pd.concat([prod,value],axis=1)
    df["日期"] = pd.to_datetime(df["日期"])

    df = df.set_index(['平台','日期','特征'])
    df = df.unstack("特征")   # 解开“特征”列
    df.columns = ["AveLimTime","AveReturn","CumulateRepay","F30Repay","F60Repay","InvestorNum","LoanNum","PlatformBackground","RegistedCapital","TradingVolume","Indicator1","Indicator2"]
    df.columns.names = ["特征"]


    # # 数据清洗
    df = df.dropna(how="all")  # 去掉缺失值
    
    return df, static_data, platforms, features

def get_curr_line_no():
    return int(pd.read_csv(CURRENT_LINE_NO_FILE_NAME).columns[0])

def write_curr_line_no(mc_row_index):
    with open(CURRENT_LINE_NO_FILE_NAME,'wt',encoding='utf-8') as f:
        f.write(str(mc_row_index)+'\n')

def get_platform_df(platform):
    plat_df = kp_df[kp_df['相关平台'].str.contains(platform)].get(['一级类','二级类','知识词典'])
    return plat_df

def parseDic(str):
    return {pair.split(': ')[0]: pair.split(': ')[1] for pair in str[1:-1].replace("'", "").split(', ')}

def get_basic_info_text(platform):
    basic_info_dict = dict(cleaned_basic_info.loc[platform])
    info_list = [f"**{k}:** {v}" for k, v in basic_info_dict.items()]
    info_text = "\n\n".join(info_list)
    return "\n\n" + info_text.strip()

def get_mc_info_text(row_index):
    mc_info_dict = dict(cleaned_mc_info.iloc[row_index])
    info_list = [f"**{k}:** {v}" for k, v in mc_info_dict.items()]
    info_text = "\n\n".join(info_list)
    return "\n\n" + info_text.strip()

def get_kp_info_text(plat_df, row_index):
    if plat_df.shape[0] == 0:
        return "～～未找到相关知识点～～"
    kp_dict = parseDic(plat_df['知识词典'].iloc[row_index])
    info_list = [f"**{k}:** {v}" for k, v in kp_dict.items()]
    info_text = "\n\n".join(info_list)
    return "\n\n" + info_text.strip()

def get_kp_row_index(row_index, page_index, page_size):
    if page_index is None:
        return row_index
    return page_index * page_size + row_index

def get_filtered_kp_df(plat_df, fc, sc):
    if fc is None and sc is None:
        return plat_df
    elif sc is None:
        return plat_df[plat_df['一级类'] == fc]
    # elif fc == '全部' and sc == '全部':
    #     return plat_df
    else:
        return plat_df[(plat_df['一级类'] == fc) & (plat_df['二级类'] == sc)]
    
def checked_mc_df_to_dict():
    '''
    Dict will be in the form:
    {mc_index: [excel_row_index, plat_index, ...]}
    '''
    return {checked_mc_df['毛刺编号'].iloc[i] : list(np.append(i+2, checked_mc_df.iloc[i][1:])) for i in range(checked_mc_df.shape[0])}

def write_row_to_excel(values, rowIndex=None):
    workBook = openpyxl.load_workbook(filename=CHECKED_MC_FILE_NAME)
    ws = workBook[SHEET_NAME]
    if rowIndex is None:
        rowIndex = ws.max_row + 1
    for col, entry in enumerate(values, start=1):
        ws.cell(row=rowIndex, column=col, value=entry)
        
    workBook.save(CHECKED_MC_FILE_NAME)
    workBook.close()
    
    return rowIndex